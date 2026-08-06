import calendar
from datetime import date
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from fieldmark.workers.models import Worker
from fieldmark.attendance.models import AttendanceRecord
from fieldmark.leave.models import LeaveRequest

def generate_payroll_excel(month_str):
    """
    Generate a styled openpyxl Excel file for worker payroll stats
    for a given month (format: 'YYYY-MM').
    """
    try:
        year, month = map(int, month_str.split('-'))
    except (ValueError, AttributeError):
        # Fallback to current month
        today = date.today()
        year, month = today.year, today.month

    # Get total calendar days and calculate working days (excluding Sundays)
    num_days = calendar.monthrange(year, month)[1]
    start_date = date(year, month, 1)
    end_date = date(year, month, num_days)
    
    total_working_days = 0
    for day in range(1, num_days + 1):
        if date(year, month, day).weekday() != 6: # 6 is Sunday
            total_working_days += 1

    wb = Workbook()
    ws = wb.active
    ws.title = f"Payroll {month_str}"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True

    # Styling elements
    primary_fill = PatternFill(start_color="3A7C3A", end_color="3A7C3A", fill_type="solid")
    header_fill = PatternFill(start_color="F0FAF0", end_color="F0FAF0", fill_type="solid")
    border_thin = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    # 1. Title Block
    ws.merge_cells("A1:H1")
    ws["A1"] = f"FieldMark Agricultural Worker Payroll Summary - {calendar.month_name[month]} {year}"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = primary_fill
    ws.row_dimensions[1].height = 40

    # 2. Table Headers
    headers = [
        "Employee ID", "Worker Name", "Zone", "Worker Type", 
        "Days Present", "Days Absent", "Leave Days", "Working Days"
    ]
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=text)
        cell.font = Font(name="Calibri", size=11, bold=True, color="3A7C3A")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
        cell.border = border_thin
    ws.row_dimensions[2].height = 25

    # 3. Populate Rows
    workers = Worker.objects.filter(is_superuser=False).order_by('name')
    row_idx = 3
    for worker in workers:
        # Days Present (Approved attendance)
        present = AttendanceRecord.objects.filter(
            worker=worker,
            date__range=[start_date, end_date],
            status=AttendanceRecord.StatusChoices.APPROVED
        ).count()

        # Approved leaves
        leave_days = 0
        leaves = LeaveRequest.objects.filter(
            worker=worker,
            status=LeaveRequest.StatusChoices.APPROVED,
            start_date__lte=end_date,
            end_date__gte=start_date
        )
        for l in leaves:
            l_start = max(l.start_date, start_date)
            l_end = min(l.end_date, end_date)
            leave_days += (l_end - l_start).days + 1

        absent = max(0, total_working_days - present - leave_days)

        row_data = [
            worker.employee_id or "N/A",
            worker.name,
            worker.assigned_zone.name if worker.assigned_zone else "N/A",
            worker.get_worker_type_display(),
            present,
            absent,
            leave_days,
            total_working_days
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=11)
            cell.border = border_thin
            if col_idx in [1, 3, 4]:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx >= 5:
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

        ws.row_dimensions[row_idx].height = 20
        row_idx += 1

    # Adjust columns widths dynamically
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Save to binary output buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=FieldMark_Payroll_{month_str}.xlsx'
    return response
